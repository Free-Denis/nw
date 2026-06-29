from django.conf import settings
from django.http import FileResponse, JsonResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt

from pathlib import Path as PyPath
import base64
import copy
import json
import shutil
import threading
import traceback
import uuid

import openpyxl
import pandas as pd

from .api_queue import API_QUEUE, ApiQueueTask


UPLOAD_BASE_DIR = PyPath("storage/uploads")
RATED_DIR = PyPath("storage/rated")
NEWS_DIR = PyPath("storage/news")
SUMMARY_DIR = PyPath("storage/summaries")
ZIP_DIR = PyPath("storage/zip")
SETTINGS_DIR = PyPath("storage/settings")
SETTINGS_FILE = SETTINGS_DIR / "ui_settings.json"
ANALYSIS_STATE_FILE = SETTINGS_DIR / "analysis_state.json"

DEFAULT_PROMPT = (
    "Оцени важность этого документа для аудиторов Сбера по шкале от 0 (неважно) "
    "до 100 (критически важно). Учитывай влияние на репутацию, регуляторные риски, "
    "финансовые последствия и операционные риски. Масштаб влияния (отдел, "
    "подразделение, компания, отрасль) учитывай в самой оценке.\n\n"
    "Ответь строго валидным JSON без markdown: "
    '{"score": 0, "reasoning": "кратко почему"}'
)

DEFAULT_SUMMARY_PROMPT = (
    "Ты — редактор новостного Telegram-канала службы внутреннего аудита Сбера. "
    "По тексту документа напиши: (1) краткое содержание — 2-3 предложения, ёмко и по делу; "
    "(2) с новой строки блок «Чем полезно аудиторам:» — 1-2 предложения о том, на какой "
    "риск или зону контроля обратить внимание. Без заголовка, без хештегов, без эмодзи."
)

DEFAULT_IMAGE_PROMPT = "Ты — Василий Кандинский, художник. Ты можешь генерировать изображения в стиле Flat Design для корпоративных презентаций Сбера. Минималистичные, плоские цвета, без мелких деталей, без текста на картинке, без людей. никаких символов, текста, букв."

DEFAULT_TEXT_MODEL = "glm-5.1"
DEFAULT_IMAGE_MODEL = "GigaChat-3-Ultra"

for directory in [UPLOAD_BASE_DIR, RATED_DIR, NEWS_DIR, SUMMARY_DIR, ZIP_DIR, SETTINGS_DIR]:
    directory.mkdir(parents=True, exist_ok=True)

ANALYSIS_JOBS = {}
JOBS_LOCK = threading.Lock()


def index(request):
    settings_data = _load_settings()
    return render(
        request,
        "posts/index.html",
        {
            "SCRIPT_NAME": settings.FORCE_SCRIPT_NAME,
            "initial_prompt": settings_data["active_prompt_text"],
            "initial_summary_prompt": settings_data["active_summary_text"],
            "initial_image_prompt": settings_data["active_image_text"],
        },
    )


def _read_json_file(path):
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def _get_default_settings():
    return {
        "active_prompt_name": "Стандартный аудит Сбера",
        "active_prompt_text": DEFAULT_PROMPT,
        "active_summary_name": "Пост в Telegram-канал",
        "active_summary_text": DEFAULT_SUMMARY_PROMPT,
        "active_image_name": "Flat Design (Корпоративный)",
        "active_image_text": DEFAULT_IMAGE_PROMPT,
        "active_text_model": DEFAULT_TEXT_MODEL,
        "active_image_model": DEFAULT_IMAGE_MODEL,
        "profiles_analysis": [
            {"name": "Стандартный аудит Сбера", "text": DEFAULT_PROMPT}
        ],
        "profiles_summary": [
            {"name": "Пост в Telegram-канал", "text": DEFAULT_SUMMARY_PROMPT},
            {"name": "Строгое резюме", "text": "Напиши строгое корпоративное резюме новости в 2 абзаца без эмодзи. Фокус на фактах и метриках."}
        ],
        "profiles_image": [
            {"name": "Flat Design (Корпоративный)", "text": DEFAULT_IMAGE_PROMPT},
            {"name": "3D Isometric", "text": "Создай иллюстрацию в стиле 3D Isometric. Яркие цвета, фокус на технологиях. Без текста и букв."}
        ]
    }


def _load_settings():
    defaults = _get_default_settings()
    if not SETTINGS_FILE.exists():
        return defaults
    try:
        data = json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
        if "profiles_analysis" not in data:
            defaults["active_prompt_text"] = data.get("prompt", DEFAULT_PROMPT)
            defaults["active_summary_text"] = data.get("summary_prompt", DEFAULT_SUMMARY_PROMPT)
            defaults["active_image_text"] = data.get("image_prompt", DEFAULT_IMAGE_PROMPT)
            return defaults
        # backfill ключей, которых не было в старых сохранениях
        data.setdefault("active_text_model", DEFAULT_TEXT_MODEL)
        data.setdefault("active_image_model", DEFAULT_IMAGE_MODEL)
        return data
    except Exception:
        return defaults


def _save_settings_to_file(data):
    SETTINGS_FILE.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _save_analysis_state(data):
    ANALYSIS_STATE_FILE.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _clear_analysis_state():
    try:
        if ANALYSIS_STATE_FILE.exists():
            ANALYSIS_STATE_FILE.unlink()
    except Exception:
        pass


def _safe_unlink(path):
    if not path:
        return
    try:
        path = PyPath(path)
        if path.exists():
            path.unlink()
    except Exception:
        pass


def _clear_directory(directory):
    if not directory.exists():
        return
    for item in directory.iterdir():
        try:
            if item.is_dir():
                shutil.rmtree(item)
            else:
                item.unlink()
        except Exception:
            pass


def _build_excel_preview(file_path):
    preview = {}
    with pd.ExcelFile(file_path) as xl:
        sheets = xl.sheet_names
        for sheet in sheets:
            df_preview = pd.read_excel(file_path, sheet_name=sheet, nrows=4)
            df_count = pd.read_excel(file_path, sheet_name=sheet, usecols=[0])
            preview[sheet] = {
                "columns": [str(column) for column in df_preview.columns],
                "total_rows": len(df_count),
            }
    return sheets, preview


def _has_running_job():
    with JOBS_LOCK:
        has_active_job = any(job["status"] == "running" for job in ANALYSIS_JOBS.values())
        has_pending_regeneration = any(
            message["text_pending"] or message["image_pending"]
            for job in ANALYSIS_JOBS.values()
            for message in job["messages"].values()
        )
    return has_active_job or has_pending_regeneration or API_QUEUE.has_pending_work()


def _job_snapshot(job):
    messages = []
    for message in sorted(job["messages"].values(), key=lambda item: item["row_index"]):
        messages.append(
            {
                "row_index": message["row_index"],
                "score": message["score"],
                "summary": message["summary"],
                "image_url": message["image_url"],
                "text_pending": message["text_pending"],
                "image_pending": message["image_pending"],
            }
        )

    resume_available = job["rated_rows"] < job["total"]
    has_pending_actions = any(
        message["text_pending"] or message["image_pending"]
        for message in job["messages"].values()
    )

    return {
        "status": job["status"],
        "stage": job["stage"],
        "current": job["current"],
        "total": job["total"],
        "rated_rows": job["rated_rows"],
        "resume_available": resume_available,
        "rated_file": job["rated_file"],
        "download_url": job["download_url"],
        "messages": messages,
        "error": job["error"],
        "should_poll": job["status"] == "running" or has_pending_actions or API_QUEUE.has_pending_work(),
        "sheet": job["sheet"],
        "column": job["column"],
        "threshold": job["threshold"],
        "prompt": job["prompt"],
        "file_path": job["file_path"],
    }


def _new_job(file_path, rated_path, sheet, column, threshold, prompt, generate_image, add_link, summary_prompt, image_prompt, text_model, image_model):
    return {
        "status": "running",
        "stage": "Запуск...",
        "current": 0,
        "total": 0,
        "rated_rows": 0,
        "rated_file": str(rated_path.absolute()),
        "download_url": f"download_rated/?path={rated_path.absolute()}",
        "error": None,
        "file_path": file_path,
        "sheet": sheet,
        "column": column,
        "threshold": threshold,
        "prompt": prompt,
        "generate_image": generate_image,
        "add_link": add_link,
        "summary_prompt": summary_prompt,
        "image_prompt": image_prompt,
        "text_model": text_model,
        "image_model": image_model,
        "messages": {},
    }


def _persist_job_state(job_id):
    with JOBS_LOCK:
        job = ANALYSIS_JOBS.get(job_id)
        if not job:
            return
        snapshot = {
            "job_id": job_id,
            "job": copy.deepcopy(job),
        }
    _save_analysis_state(snapshot)


def _normalize_persisted_job(data):
    if not data or "job" not in data:
        return None

    payload = copy.deepcopy(data)
    job = payload.get("job", {})
    file_path = job.get("file_path", "")
    sheet = job.get("sheet", "")
    rated_file = job.get("rated_file", "")

    if not file_path or not sheet:
        return None
    if not PyPath(file_path).exists():
        return None

    try:
        source_df = pd.read_excel(file_path, sheet_name=sheet)
        total_rows = len(source_df)
    except Exception:
        return None

    rated_rows = 0
    if rated_file and PyPath(rated_file).exists():
        try:
            rated_df = pd.read_excel(rated_file, sheet_name=str(sheet)[:31])
            rated_rows = len(rated_df)
        except Exception:
            rated_rows = 0

    job["total"] = total_rows
    job["rated_rows"] = rated_rows
    job["current"] = min(job.get("current", rated_rows), total_rows)

    if job.get("status") == "running":
        job["status"] = "interrupted"
        job["stage"] = "Сервер был перезапущен. Можно продолжить с первой необработанной строки."

    for message in job.get("messages", {}).values():
        message["text_pending"] = False
        message["image_pending"] = False

    if "messages" in job:
        job["messages"] = {int(k): v for k, v in job["messages"].items()}

    return payload


def _load_persisted_state():
    return _normalize_persisted_job(_read_json_file(ANALYSIS_STATE_FILE))


def _safe_name(value):
    cleaned = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in str(value))
    cleaned = cleaned.strip("_")
    return cleaned[:64] or "sheet"


def _rated_path_for(sheet):
    return RATED_DIR / f"rated_{_safe_name(sheet)}.xlsx"


def _summary_path_for(sheet, row_index):
    return SUMMARY_DIR / f"{_safe_name(sheet)}_{row_index}.txt"


def _serialize_cell(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, (int, float, bool)):
        return value
    return str(value)


def _initialize_rated_workbook(rated_path, source_columns, sheet_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    headers = list(source_columns)
    if "Оценка" not in headers:
        headers.append("Оценка")
    if "Обоснование" not in headers:
        headers.append("Обоснование")
    ws.append(headers)
    wb.save(rated_path)
    return headers


def _append_processed_row(rated_path, sheet_name, headers, row_dict):
    wb = openpyxl.load_workbook(rated_path)
    ws = wb[sheet_name[:31]]
    ws.append([_serialize_cell(row_dict.get(header)) for header in headers])
    wb.save(rated_path)


def _update_score_in_excel(rated_path, sheet_name, row_index, new_score):
    wb = openpyxl.load_workbook(rated_path)
    ws = wb[sheet_name[:31]]
    header_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    score_col = header_map.get("Оценка")
    if not score_col:
        wb.close()
        raise ValueError("В Excel не найден столбец 'Оценка'")
    excel_row = row_index + 2
    ws.cell(row=excel_row, column=score_col).value = new_score
    wb.save(rated_path)


def _save_summary(sheet, row_index, summary):
    summary_path = _summary_path_for(sheet, row_index)
    summary_path.write_text(summary, encoding="utf-8")
    return summary_path


def _save_image(sheet, row_index, image_data, old_filename=None):
    if not image_data or not image_data.startswith("data:image"):
        return None, None

    if old_filename:
        _safe_unlink(NEWS_DIR / old_filename)

    format_part = image_data.split(";")[0]
    ext = "jpg" if ("jpeg" in format_part or "jpg" in format_part) else "png"
    b64_data = image_data.split(",", 1)[1]
    image_bytes = base64.b64decode(b64_data)
    filename = f"{_safe_name(sheet)}_{row_index}_{uuid.uuid4().hex[:8]}.{ext}"
    target_path = NEWS_DIR / filename
    with open(target_path, "wb") as output_file:
        output_file.write(image_bytes)
    return filename, f"news_image/?file={filename}"


def _upsert_message(job_id, row_index, score, original_text, summary, summary_path, image_file, image_url):
    with JOBS_LOCK:
        job = ANALYSIS_JOBS.get(job_id)
        if not job:
            return
        current = job["messages"].get(row_index, {})
        job["messages"][row_index] = {
            "row_index": row_index,
            "score": score,
            "original_text": original_text,
            "summary": summary,
            "summary_path": str(summary_path) if summary_path else current.get("summary_path"),
            "image_file": image_file,
            "image_url": image_url,
            "text_pending": current.get("text_pending", False),
            "image_pending": current.get("image_pending", False),
        }
    _persist_job_state(job_id)


def _remove_message(job_id, row_index):
    with JOBS_LOCK:
        job = ANALYSIS_JOBS.get(job_id)
        if not job:
            return None
        removed = job["messages"].pop(row_index, None)
    _persist_job_state(job_id)
    return removed


def _set_message_pending(job_id, row_index, action, pending):
    with JOBS_LOCK:
        job = ANALYSIS_JOBS.get(job_id)
        if not job or row_index not in job["messages"]:
            return
        if action == "text":
            job["messages"][row_index]["text_pending"] = pending
        elif action == "image":
            job["messages"][row_index]["image_pending"] = pending
    _persist_job_state(job_id)


def _set_job_status(job_id, **updates):
    with JOBS_LOCK:
        job = ANALYSIS_JOBS.get(job_id)
        if not job:
            return
        job.update(updates)
    _persist_job_state(job_id)


def _build_retry_callback(job_id, label, row_number):
    def _callback(_exc, attempt, max_retries):
        _set_job_status(
            job_id,
            stage=f"{label}: ошибка на строке {row_number}, повтор через 5 сек ({attempt + 1}/{max_retries})",
        )

    return _callback


def _queue_task_and_wait(task):
    API_QUEUE.submit(task)
    task.done.wait()
    return task.result, task.error


def _run_analysis_worker(job_id):
    try:
        from .gigachat import GigaChatClient

        with JOBS_LOCK:
            job = ANALYSIS_JOBS[job_id]
            file_path = job["file_path"]
            sheet = job["sheet"]
            column = job["column"]
            threshold = job["threshold"]
            prompt = job["prompt"]
            summary_prompt = job.get("summary_prompt", DEFAULT_SUMMARY_PROMPT)
            image_prompt = job.get("image_prompt", DEFAULT_IMAGE_PROMPT)
            text_model = job.get("text_model", DEFAULT_TEXT_MODEL)
            image_model = job.get("image_model", DEFAULT_IMAGE_MODEL)
            generate_image_flag = job.get("generate_image", True)
            add_link_flag = job.get("add_link", True)
            rated_path = PyPath(job["rated_file"])

        gc = GigaChatClient.get_instance()
        source_df = pd.read_excel(file_path, sheet_name=sheet)

        if column not in source_df.columns:
            _set_job_status(job_id, status="error", error=f"Столбец '{column}' не найден")
            return

        total_rows = len(source_df)
        sheet_name = str(sheet)[:31]

        if rated_path.exists():
            rated_df = pd.read_excel(rated_path, sheet_name=sheet_name)
            processed_rows = len(rated_df)
            headers = list(rated_df.columns)
        else:
            processed_rows = 0
            headers = _initialize_rated_workbook(rated_path, source_df.columns, sheet_name)

        _set_job_status(
            job_id,
            total=total_rows,
            rated_rows=processed_rows,
            current=processed_rows,
            stage="Продолжение с первой необработанной строки" if processed_rows else "Подготовка к анализу",
            download_url=f"download_rated/?path={rated_path.absolute()}",
        )

        for position in range(processed_rows, total_rows):
            with JOBS_LOCK:
                if job_id not in ANALYSIS_JOBS:
                    return

            row = source_df.iloc[position]
            text = str(row.get(column, "")) if pd.notna(row.get(column, "")) else ""
            display_row = position + 1
            score = None
            reasoning = ""
            summary = None
            summary_path = None
            image_file = None
            image_url = None

            _set_job_status(job_id, current=display_row, stage=f"Оценка новости ({display_row}/{total_rows})")

            if text.strip() and text.strip().lower() != "nan":
                rate_task = ApiQueueTask(
                    name=f"rate-{display_row}",
                    func=lambda text=text, prompt=prompt, m=text_model: gc.rate_importance(text, prompt, model=m),
                    on_attempt_error=_build_retry_callback(job_id, "Оценка", display_row),
                )
                result, error = _queue_task_and_wait(rate_task)
                if error:
                    reasoning = f"Ошибка оценки после 3 попыток: {error}"
                elif result:
                    score, reasoning = result

            if score is not None and score >= threshold:
                _set_job_status(job_id, current=display_row, stage=f"Суммаризация новости ({display_row}/{total_rows})")

                summarize_task = ApiQueueTask(
                    name=f"summary-{display_row}",
                    func=lambda text=text, p=summary_prompt, m=text_model: gc.summarize_news(text, p, model=m),
                    on_attempt_error=_build_retry_callback(job_id, "Суммаризация", display_row),
                )
                summary_result, summary_error = _queue_task_and_wait(summarize_task)
                if summary_error:
                    reasoning = f"{reasoning}\nОшибка суммаризации: {summary_error}".strip()
                else:
                    summary = summary_result
                    
                    try:
                        # ищем колонки по подстроке (имена в выгрузках различаются)
                        type_col = next((c for c in source_df.columns if "вид документа" in str(c).lower()), None)
                        num_col = next((c for c in source_df.columns if "номер" in str(c).lower()), None)
                        subj_col = next((c for c in source_df.columns if "содержание" in str(c).lower()), None)
                        link_col = next((c for c in source_df.columns if "ссылка" in str(c).lower()), None)

                        def _get_val(c):
                            if not c: return ""
                            v = row.get(c)
                            if pd.isna(v): return ""
                            v = str(v).strip()
                            return "" if v.lower() == "nan" else v

                        assembled_text = f"{_get_val(type_col)} {_get_val(num_col)}".strip()
                        link_val = _get_val(link_col)
                        is_valid_link = link_val.startswith("http://") or link_val.startswith("https://")
                        # подпись ссылки: вид+номер → содержание → запасной текст
                        label = assembled_text or _get_val(subj_col)[:80] or "Открыть документ"

                        if is_valid_link and add_link_flag:
                            summary += f'\n\n<a href="{link_val}" target="_blank" style="color:var(--tg-primary);text-decoration:none;">{label}</a>'
                        elif assembled_text:
                            summary += f'\n\n{assembled_text}'
                    except Exception as e:
                        traceback.print_exc()

                    summary_path = _save_summary(sheet, position, summary)

                if summary and generate_image_flag:
                    _set_job_status(job_id, current=display_row, stage=f"Генерация изображения ({display_row}/{total_rows})")
                    image_task = ApiQueueTask(
                        name=f"image-{display_row}",
                        func=lambda s=summary_result, p=image_prompt, m=image_model: gc.generate_image(s, p, model=m),
                        on_attempt_error=_build_retry_callback(job_id, "Изображение", display_row),
                    )
                    image_result, _image_error = _queue_task_and_wait(image_task)
                    if image_result:
                        image_file, image_url = _save_image(sheet, position, image_result)

                if summary:
                    _upsert_message(
                        job_id=job_id,
                        row_index=position,
                        score=score,
                        original_text=text,
                        summary=summary,
                        summary_path=summary_path,
                        image_file=image_file,
                        image_url=image_url,
                    )

            row_dict = row.to_dict()
            row_dict["Оценка"] = score
            row_dict["Обоснование"] = reasoning
            _append_processed_row(rated_path, sheet_name, headers, row_dict)

            _set_job_status(job_id, rated_rows=display_row)

        _set_job_status(
            job_id,
            status="done",
            stage="Готово!",
            current=total_rows,
            rated_rows=total_rows,
            error=None,
        )
    except Exception as exc:
        traceback.print_exc()
        with JOBS_LOCK:
            job = ANALYSIS_JOBS.get(job_id)
            rated_rows = job["rated_rows"] if job else 0
        _set_job_status(
            job_id,
            status="error",
            stage="Процесс остановлен",
            current=rated_rows,
            error=str(exc),
        )


@csrf_exempt
def upload_excel(request):
    if request.method != "POST":
        return JsonResponse({"error": "Только POST"}, status=405)

    files = request.FILES.getlist("file")
    if not files:
        return JsonResponse({"error": "Файл не найден"}, status=400)

    file_obj = files[0]
    if not file_obj.name.endswith((".xlsx", ".xls")):
        return JsonResponse({"error": "Только Excel файлы"}, status=400)

    for directory in [UPLOAD_BASE_DIR, RATED_DIR, NEWS_DIR, SUMMARY_DIR, ZIP_DIR]:
        _clear_directory(directory)
    _clear_analysis_state()
    API_QUEUE.clear()

    with JOBS_LOCK:
        ANALYSIS_JOBS.clear()

    target_path = UPLOAD_BASE_DIR / file_obj.name
    with open(target_path, "wb+") as destination:
        for chunk in file_obj.chunks():
            destination.write(chunk)

    try:
        sheets, preview = _build_excel_preview(target_path)

        return JsonResponse(
            {
                "status": "ok",
                "file_path": str(target_path.absolute()),
                "sheets": sheets,
                "preview": preview,
            }
        )
    except Exception as exc:
        return JsonResponse({"error": str(exc)}, status=500)


@csrf_exempt
def save_settings(request):
    if request.method != "POST":
        return JsonResponse({"error": "Только POST"}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Невалидный JSON"}, status=400)

    _save_settings_to_file(data)
    return JsonResponse({
        "status": "ok",
        "settings": data
    })


def list_models(request):
    """Список доступных моделей нейросети (для выпадающих списков в настройках)."""
    from .gigachat import GigaChatClient
    try:
        models = GigaChatClient.get_instance().list_models()
        return JsonResponse({"status": "ok", "models": models})
    except Exception as exc:
        return JsonResponse({"status": "error", "error": str(exc), "models": []})


def app_state(request):
    with JOBS_LOCK:
        active_job_id = next(
            (job_id for job_id, job in reversed(list(ANALYSIS_JOBS.items()))),
            None,
        )

    if active_job_id:
        with JOBS_LOCK:
            job = ANALYSIS_JOBS.get(active_job_id)
            if not job:
                active_job_id = None
            else:
                payload = _job_snapshot(job)
                file_path = job["file_path"]
        if active_job_id:
            try:
                sheets, preview = _build_excel_preview(file_path)
            except Exception:
                return JsonResponse({"status": "empty", "settings": _load_settings()})
            return JsonResponse(
                {
                    "status": "ok",
                    "job_id": active_job_id,
                    "file_path": file_path,
                    "sheets": sheets,
                    "preview": preview,
                    "job": payload,
                    "settings": _load_settings(),
                }
            )

    persisted_state = _load_persisted_state()
    if not persisted_state:
        return JsonResponse({"status": "empty", "settings": _load_settings()})

    job = persisted_state["job"]
    file_path = job["file_path"]
    try:
        sheets, preview = _build_excel_preview(file_path)
    except Exception:
        return JsonResponse({"status": "empty", "settings": _load_settings()})

    persisted_snapshot = _job_snapshot(job)
    if job["status"] == "interrupted":
        persisted_snapshot["should_poll"] = False

    hydrated_job_id = persisted_state.get("job_id") or str(uuid.uuid4())
    with JOBS_LOCK:
        ANALYSIS_JOBS[hydrated_job_id] = copy.deepcopy(job)

    return JsonResponse(
        {
            "status": "ok",
            "job_id": hydrated_job_id,
            "file_path": file_path,
            "sheets": sheets,
            "preview": preview,
            "job": persisted_snapshot,
            "settings": _load_settings(),
        }
    )


@csrf_exempt
def start_analysis(request):
    if request.method != "POST":
        return JsonResponse({"error": "Только POST"}, status=405)

    if _has_running_job():
        return JsonResponse({"error": "Анализ уже выполняется"}, status=409)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Невалидный JSON"}, status=400)

    file_path = data.get("file_path", "")
    sheet = data.get("sheet", "")
    column = data.get("column", "")
    
    settings_data = _load_settings()
    final_prompt = str(data.get("prompt", "")).strip() or settings_data.get("active_prompt_text", DEFAULT_PROMPT)

    threshold = float(data.get("threshold", 70.0))
    generate_image = bool(data.get("generate_image", True))
    add_link = bool(data.get("add_link", True))

    if not file_path or not sheet or not column:
        return JsonResponse({"error": "Не заполнены обязательные поля"}, status=400)

    source_df = pd.read_excel(file_path, sheet_name=sheet)
    rated_path = _rated_path_for(sheet)
    continue_mode = False

    if rated_path.exists():
        try:
            rated_df = pd.read_excel(rated_path, sheet_name=str(sheet)[:31])
            continue_mode = len(rated_df) < len(source_df)
        except Exception:
            continue_mode = False

    if not continue_mode:
        for directory in [RATED_DIR, NEWS_DIR, SUMMARY_DIR, ZIP_DIR]:
            _clear_directory(directory)

    preserved_messages = {}
    persisted_state = _load_persisted_state()
    with JOBS_LOCK:
        if continue_mode:
            for existing_job in reversed(list(ANALYSIS_JOBS.values())):
                same_sheet = existing_job.get("sheet") == sheet
                same_rated_file = existing_job.get("rated_file") == str(rated_path.absolute())
                if same_sheet and same_rated_file and existing_job.get("messages"):
                    preserved_messages = copy.deepcopy(existing_job["messages"])
                    break
            if not preserved_messages and persisted_state:
                persisted_job = persisted_state.get("job", {})
                same_sheet = persisted_job.get("sheet") == sheet
                same_rated_file = persisted_job.get("rated_file") == str(rated_path.absolute())
                if same_sheet and same_rated_file and persisted_job.get("messages"):
                    preserved_messages = copy.deepcopy(persisted_job["messages"])
        else:
            ANALYSIS_JOBS.clear()

    job_id = str(uuid.uuid4())
    job = _new_job(
        file_path=file_path,
        rated_path=_rated_path_for(sheet),
        sheet=sheet,
        column=column,
        threshold=threshold,
        prompt=final_prompt,
        generate_image=generate_image,
        add_link=add_link,
        summary_prompt=settings_data.get("active_summary_text", DEFAULT_SUMMARY_PROMPT),
        image_prompt=settings_data.get("active_image_text", DEFAULT_IMAGE_PROMPT),
        text_model=settings_data.get("active_text_model", DEFAULT_TEXT_MODEL),
        image_model=settings_data.get("active_image_model", DEFAULT_IMAGE_MODEL),
    )
    job["messages"] = preserved_messages

    with JOBS_LOCK:
        ANALYSIS_JOBS[job_id] = job
    _persist_job_state(job_id)

    worker = threading.Thread(target=_run_analysis_worker, args=(job_id,), daemon=True)
    worker.start()
    return JsonResponse({"job_id": job_id, "continue_mode": continue_mode})


def analysis_progress(request):
    job_id = request.GET.get("job_id", "")
    with JOBS_LOCK:
        job = ANALYSIS_JOBS.get(job_id)
        if not job:
            return JsonResponse({"error": "Job not found"}, status=404)
        payload = _job_snapshot(job)
    return JsonResponse(payload)


@csrf_exempt
def lock_message(request):
    if request.method != "POST":
        return JsonResponse({"error": "Только POST"}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Невалидный JSON"}, status=400)

    job_id = data.get("job_id", "")
    row_index = int(data.get("row_index"))

    with JOBS_LOCK:
        job = ANALYSIS_JOBS.get(job_id)
        if not job:
            return JsonResponse({"error": "Job not found"}, status=404)
        message = job["messages"].get(row_index)
        if not message:
            return JsonResponse({"error": "Сообщение не найдено"}, status=404)

    try:
        _update_score_in_excel(PyPath(job["rated_file"]), job["sheet"], row_index, 1)
        _safe_unlink(message.get("summary_path"))
        if message.get("image_file"):
            _safe_unlink(NEWS_DIR / message["image_file"])
        _remove_message(job_id, row_index)
        _set_job_status(job_id, stage=f"Сообщение по строке {row_index + 1} заблокировано")
        return JsonResponse({"status": "ok"})
    except Exception as exc:
        return JsonResponse({"error": str(exc)}, status=500)


@csrf_exempt
def regenerate_text(request):
    if request.method != "POST":
        return JsonResponse({"error": "Только POST"}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Невалидный JSON"}, status=400)

    job_id = data.get("job_id", "")
    row_index = int(data.get("row_index"))

    with JOBS_LOCK:
        job = ANALYSIS_JOBS.get(job_id)
        if not job:
            return JsonResponse({"error": "Job not found"}, status=404)
        message = job["messages"].get(row_index)
        if not message:
            return JsonResponse({"error": "Сообщение не найдено"}, status=404)
        if message["text_pending"]:
            return JsonResponse({"error": "Перегенерация текста уже в очереди"}, status=409)
        original_text = message["original_text"]

    from .gigachat import GigaChatClient

    gc = GigaChatClient.get_instance()
    _set_message_pending(job_id, row_index, "text", True)
    _set_job_status(job_id, stage=f"Перегенерация текста для строки {row_index + 1}")

    def _on_success(summary):
        with JOBS_LOCK:
            current_job = ANALYSIS_JOBS.get(job_id)
            current_message = current_job["messages"].get(row_index) if current_job else None
            if not current_message:
                return
            summary_path = _save_summary(current_job["sheet"], row_index, summary)
            current_message["summary"] = summary
            current_message["summary_path"] = str(summary_path)
            current_message["text_pending"] = False
        _set_job_status(job_id, stage=f"Текст для строки {row_index + 1} обновлён")

    def _on_failure(_exc):
        _set_message_pending(job_id, row_index, "text", False)
        _set_job_status(job_id, stage=f"Не удалось обновить текст для строки {row_index + 1}")

    task = ApiQueueTask(
        name=f"regen-text-{row_index}",
        priority=True,
        func=lambda p=job.get("summary_prompt", DEFAULT_SUMMARY_PROMPT), m=job.get("text_model", DEFAULT_TEXT_MODEL): gc.summarize_news(original_text, p, model=m),
        on_attempt_error=_build_retry_callback(job_id, "Перегенерация текста", row_index + 1),
        on_success=_on_success,
        on_failure=_on_failure,
    )
    API_QUEUE.submit(task)
    return JsonResponse({"status": "queued"})


@csrf_exempt
def regenerate_image(request):
    if request.method != "POST":
        return JsonResponse({"error": "Только POST"}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Невалидный JSON"}, status=400)

    job_id = data.get("job_id", "")
    row_index = int(data.get("row_index"))

    with JOBS_LOCK:
        job = ANALYSIS_JOBS.get(job_id)
        if not job:
            return JsonResponse({"error": "Job not found"}, status=404)
        message = job["messages"].get(row_index)
        if not message:
            return JsonResponse({"error": "Сообщение не найдено"}, status=404)
        if message["image_pending"]:
            return JsonResponse({"error": "Перегенерация изображения уже в очереди"}, status=409)
        summary = message["summary"]
        image_file = message.get("image_file")

    from .gigachat import GigaChatClient

    gc = GigaChatClient.get_instance()
    _set_message_pending(job_id, row_index, "image", True)
    _set_job_status(job_id, stage=f"Перегенерация изображения для строки {row_index + 1}")

    def _on_success(image_data):
        with JOBS_LOCK:
            current_job = ANALYSIS_JOBS.get(job_id)
            current_message = current_job["messages"].get(row_index) if current_job else None
            if not current_message:
                return
            new_file, new_url = _save_image(
                current_job["sheet"],
                row_index,
                image_data,
                old_filename=current_message.get("image_file"),
            )
            current_message["image_file"] = new_file
            current_message["image_url"] = new_url
            current_message["image_pending"] = False
        _set_job_status(job_id, stage=f"Изображение для строки {row_index + 1} обновлено")

    def _on_failure(_exc):
        _set_message_pending(job_id, row_index, "image", False)
        _set_job_status(job_id, stage=f"Не удалось обновить изображение для строки {row_index + 1}")

    task = ApiQueueTask(
        name=f"regen-image-{row_index}",
        priority=True,
        func=lambda p=job.get("image_prompt", DEFAULT_IMAGE_PROMPT), m=job.get("image_model", DEFAULT_IMAGE_MODEL): gc.generate_image(summary, p, model=m),
        on_attempt_error=_build_retry_callback(job_id, "Перегенерация изображения", row_index + 1),
        on_success=_on_success,
        on_failure=_on_failure,
    )
    API_QUEUE.submit(task)
    return JsonResponse({"status": "queued"})


def download_rated(request):
    file_path = request.GET.get("path")
    if not file_path:
        return JsonResponse({"error": "No path"}, status=400)
    path = PyPath(file_path)
    if not path.exists():
        return JsonResponse({"error": "File not found"}, status=404)
    try:
        return FileResponse(open(path, "rb"), as_attachment=True, filename=path.name)
    except Exception as exc:
        return JsonResponse({"error": str(exc)}, status=500)


def news_image(request):
    filename = request.GET.get("file")
    if not filename:
        return JsonResponse({"error": "No file"}, status=400)
    path = NEWS_DIR / filename
    if not path.exists():
        return JsonResponse({"error": "File not found"}, status=404)
    try:
        return FileResponse(open(path, "rb"), as_attachment=False, filename=path.name)
    except Exception as exc:
        return JsonResponse({"error": str(exc)}, status=500)


# ══════════════════════════════════════════════════════════════════
# Prompt Laboratory endpoints
# ══════════════════════════════════════════════════════════════════

from .prompt_lab import LAB_JOBS, LAB_LOCK, lab_job_snapshot, run_analysis_step, run_generation_step, _new_lab_job, _update_lab_job


@csrf_exempt
def lab_upload_excel(request):
    """Handle excel upload for prompt lab and return preview data (up to 20 rows per sheet)."""
    if request.method != "POST":
        return JsonResponse({"error": "Только POST"}, status=405)

    if "file" not in request.FILES:
        return JsonResponse({"error": "Нет файла"}, status=400)

    file_obj = request.FILES["file"]
    filename = file_obj.name

    if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
        return JsonResponse({"error": "Неверный формат файла"}, status=400)

    file_path = UPLOAD_BASE_DIR / f"lab_upload_{uuid.uuid4().hex}.xlsx"
    with open(file_path, "wb") as f:
        for chunk in file_obj.chunks():
            f.write(chunk)

    try:
        lab_excel_data = []
        with pd.ExcelFile(file_path) as xl:
            sheets = xl.sheet_names
            for sheet in sheets:
                df = pd.read_excel(file_path, sheet_name=sheet, nrows=40).fillna("")
                columns = [str(c) for c in df.columns]
                data = df.values.tolist()
                lab_excel_data.append({"sheet": sheet, "columns": columns, "data": data})
        
        # We can safely delete the file as we only need the preview data for labeling
        _safe_unlink(file_path)
        return JsonResponse({"excel_data": lab_excel_data})
    except Exception as exc:
        _safe_unlink(file_path)
        return JsonResponse({"error": str(exc)}, status=500)



@csrf_exempt
def lab_start(request):
    """Step 1 → Step 2: accept examples + user input, start background analysis."""
    if request.method != "POST":
        return JsonResponse({"error": "Только POST"}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Невалидный JSON"}, status=400)

    examples = data.get("examples", [])
    if len(examples) < 2:
        return JsonResponse({"error": "Нужно минимум 2 примера"}, status=400)

    has_positive = any(e.get("label") == "positive" for e in examples)
    if not has_positive:
        return JsonResponse({"error": "Нужен хотя бы 1 позитивный пример"}, status=400)

    has_negative = any(e.get("label") == "negative" for e in examples)
    if not has_negative:
        return JsonResponse({"error": "Нужен хотя бы 1 негативный пример, иначе лаборатория не поймёт границу важности"}, status=400)

    user_comment = str(data.get("user_comment", "")).strip()
    if len(user_comment) < 10:
        return JsonResponse({"error": "Опишите задачу подробнее (минимум 10 символов)"}, status=400)

    user_input = {
        "examples": examples,
        "user_comment": user_comment,
        "target_audience": str(data.get("target_audience", "")).strip(),
        "success_threshold": int(data.get("success_threshold", 70)),
        "scale": str(data.get("scale", "")).strip(),
        "topics": data.get("topics", []),
        "ask_questions": bool(data.get("ask_questions", True)),
    }

    job_id = str(uuid.uuid4())
    job = _new_lab_job(user_input)
    job["text_model"] = _load_settings().get("active_text_model", DEFAULT_TEXT_MODEL)

    with LAB_LOCK:
        LAB_JOBS[job_id] = job

    worker = threading.Thread(target=run_analysis_step, args=(job_id,), daemon=True)
    worker.start()

    return JsonResponse({"job_id": job_id})


def lab_progress(request):
    """Poll current lab job status."""
    job_id = request.GET.get("job_id", "")
    with LAB_LOCK:
        job = LAB_JOBS.get(job_id)
        if not job:
            return JsonResponse({"error": "Job not found"}, status=404)
        snapshot = lab_job_snapshot(job)
    return JsonResponse(snapshot)


@csrf_exempt
def lab_cancel(request):
    if request.method != "POST":
        return JsonResponse({"error": "Только POST"}, status=405)
    try:
        data = json.loads(request.body)
        job_id = data.get("job_id")
        with LAB_LOCK:
            job = LAB_JOBS.get(job_id)
            if job:
                job["cancelled"] = True
                job["status"] = "error"
                job["error"] = "Процесс отменен пользователем."
        return JsonResponse({"ok": True})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def lab_answer_questions(request):
    """Step 3 → Step 4: accept user answers, start prompt generation."""
    if request.method != "POST":
        return JsonResponse({"error": "Только POST"}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Невалидный JSON"}, status=400)

    job_id = data.get("job_id", "")
    answers = data.get("answers", {})

    with LAB_LOCK:
        job = LAB_JOBS.get(job_id)
        if not job:
            return JsonResponse({"error": "Job not found"}, status=404)
        if job["status"] != "questions_ready":
            return JsonResponse({"error": f"Неверный статус: {job['status']}"}, status=409)
        job["answers"] = answers
        job["status"] = "generating"
        job["step"] = 4

    worker = threading.Thread(target=run_generation_step, args=(job_id,), daemon=True)
    worker.start()

    return JsonResponse({"status": "ok"})


@csrf_exempt
def lab_iterate(request):
    """Step 4: run another improvement iteration."""
    if request.method != "POST":
        return JsonResponse({"error": "Только POST"}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Невалидный JSON"}, status=400)

    job_id = data.get("job_id", "")

    with LAB_LOCK:
        job = LAB_JOBS.get(job_id)
        if not job:
            return JsonResponse({"error": "Job not found"}, status=404)
        if job["status"] != "done":
            return JsonResponse({"error": f"Неверный статус: {job['status']}"}, status=409)
        job["status"] = "generating"

    worker = threading.Thread(target=run_generation_step, args=(job_id, True), daemon=True)
    worker.start()

    return JsonResponse({"status": "ok"})


@csrf_exempt
def lab_save_profile(request):
    """Save generated prompt as a profile in settings."""
    if request.method != "POST":
        return JsonResponse({"error": "Только POST"}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Невалидный JSON"}, status=400)

    profile_name = str(data.get("profile_name", "")).strip()
    prompt_text = str(data.get("prompt_text", "")).strip()

    if not profile_name or not prompt_text:
        return JsonResponse({"error": "Имя профиля и текст промпта обязательны"}, status=400)

    settings_data = _load_settings()
    profiles = settings_data.get("profiles_analysis", [])

    existing = next((p for p in profiles if p["name"] == profile_name), None)
    if existing:
        existing["text"] = prompt_text
    else:
        profiles.append({"name": profile_name, "text": prompt_text})

    settings_data["profiles_analysis"] = profiles
    settings_data["active_prompt_name"] = profile_name
    settings_data["active_prompt_text"] = prompt_text
    _save_settings_to_file(settings_data)

    return JsonResponse({"status": "ok", "settings": settings_data})
